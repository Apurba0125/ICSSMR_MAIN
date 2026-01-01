
from django.contrib import admin
from django.http import HttpResponse
import openpyxl


admin.site.site_header = "ICSSMR  Admin"
admin.site.site_title = "ICSSMR Admin Portal"
admin.site.index_title = "Welcome to ICSSMR Admin"

from .models import StudentRegistration


def download_excel(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    headers = [
        "Name", "DOB", "Gender", "Email", "Mobile", "WhatsApp",
        "Qualification", "Course",
        "Address Line 1", "Address Line 2",
        "City", "District", "State", "Country", "Pin",
        "Marks 10", "Marks 12", "Marks UG", "Marks PG"
    ]
    ws.append(headers)

    for obj in queryset:
        ws.append([
            obj.name,
            obj.dob,
            obj.gender,
            obj.email,
            obj.mobile,
            obj.whatsapp,
            obj.qualification,
            obj.course,
            obj.address1,
            obj.address2,
            obj.city,
            obj.district,
            obj.state,
            obj.country,
            obj.pin,
            obj.marks_10,
            obj.marks_12,
            obj.marks_ug,
            obj.marks_pg,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename=students.xlsx'

    wb.save(response)
    return response


download_excel.short_description = "⬇ Download selected students as Excel"



@admin.register(StudentRegistration)
class StudentRegistrationAdmin(admin.ModelAdmin):

    list_display = (
        'name',
        'mobile',
        'email',
        'course',
        'qualification',
        'created_at',
    )

    search_fields = ('name', 'mobile', 'email', 'course')
    list_filter = ('qualification', 'course', 'state')

    actions = [download_excel]

    fieldsets = (
        ('Personal Information', {
            'fields': ('name', 'dob', 'gender')
        }),
        ('Contact Information', {
            'fields': ('email', 'mobile', 'whatsapp')
        }),
        ('Qualification & Course', {
            'fields': ('qualification', 'course')
        }),
        ('Address Details', {
            'fields': (
                'address1',
                'address2',
                'city',
                'district',
                'state',
                'country',
                'pin'
            )
        }),
        ('Academic Marks', {
            'fields': (
                'marks_10',
                'marks_12',
                'marks_ug',
                'marks_pg'
            )
        }),
        ('Uploaded Documents', {
            'fields': (
                'ug_marksheet',
                'x_admit',
                'x_marksheet',
                'xii_marksheet',
                'aadhar',
                'cv',
                'photo'
            )
        }),
    )





# associate membership 
from django.contrib import admin
from django.http import HttpResponse
from openpyxl import Workbook
from .models import AssociateMembership


def export_as_excel(modeladmin, request, queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Associate Membership"

    # ---------- Header Row ----------
    headers = [
        "Name", "Mobile", "Email", "WhatsApp",
        "Salutation", "DOB", "Gender",
        "Organization", "Designation", "Experience",
        "Sector", "Qualification",
        "Address Line 1", "Address Line 2",
        "City", "District", "State", "Country", "Pin",
        "Created At"
    ]

    ws.append(headers)

    # ---------- Data Rows ----------
    for obj in queryset:
        ws.append([
            obj.name,
            obj.mobile,
            obj.email,
            obj.whatsapp,
            obj.salutation,
            obj.dob,
            obj.gender,
            obj.organization,
            obj.designation,
            obj.experience,
            obj.sector,
            obj.qualification,
            obj.address1,
            obj.address2,
            obj.city,
            obj.district,
            obj.state,
            obj.country,
            obj.pin,
            obj.created_at.strftime("%Y-%m-%d %H:%M")
        ])

    # ---------- Response ----------
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=associate_membership.xlsx'

    wb.save(response)
    return response


export_as_excel.short_description = "⬇️ Download selected as Excel"


@admin.register(AssociateMembership)
class AssociateMembershipAdmin(admin.ModelAdmin):
    list_display = (
        'name',
        'mobile',
        'email',
        'organization',
        'designation',
        'sector',
        'qualification',
        'created_at'
    )

    search_fields = ('name', 'mobile', 'email', 'organization')
    list_filter = ('sector', 'qualification', 'state')

    actions = [export_as_excel]





# faculty registration
from django.contrib import admin
from django.http import HttpResponse
from openpyxl import Workbook

from .models import FacultyMembership


def export_faculty_as_excel(modeladmin, request, queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Faculty Membership Data"

    # Excel Header Row
    ws.append([
        "Name",
        "Mobile",
        "Email",
        "Institution",
        "Designation",
        "Domain",
        "Specialization",
        "Qualification",
        "Experience",
        "Research Interest",
        "City",
        "District",
        "State",
        "Country",
        "Pin",
        "Created At",
    ])

    # Data Rows
    for obj in queryset:
        ws.append([
            obj.name,
            obj.mobile,
            obj.email,
            obj.institution,
            obj.designation,
            obj.domain,
            obj.specialization,
            obj.qualification,
            obj.experience,
            obj.research_interest,
            obj.city,
            obj.district,
            obj.state,
            obj.country,
            obj.pin,
            obj.created_at.strftime("%d-%m-%Y %H:%M"),
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Faculty_Membership.xlsx"'

    wb.save(response)
    return response


export_faculty_as_excel.short_description = "📥 Download selected records as Excel"


@admin.register(FacultyMembership)
class FacultyMembershipAdmin(admin.ModelAdmin):

    list_display = (
        "name",
        "mobile",
        "email",
        "institution",
        "designation",
        "domain",
        "qualification",
        "created_at",
    )

    search_fields = ("name", "mobile", "email", "institution")
    list_filter = ("domain", "designation", "qualification", "state")
    readonly_fields = ("created_at",)

    actions = [export_faculty_as_excel]



#student membership
from django.contrib import admin
from django.http import HttpResponse
from .models import StudentMembership
import openpyxl
from openpyxl.styles import Font


@admin.register(StudentMembership)
class StudentMembershipAdmin(admin.ModelAdmin):
    list_display = (
        'name', 'mobile', 'email',
        'qualification', 'internship',
        'created_at'
    )
    search_fields = ('name', 'mobile', 'email')
    list_filter = ('qualification', 'internship', 'created_at')

    actions = ['download_excel']

    def download_excel(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Student Membership Data"

        # ===== Excel Headers (ALL DATA) =====
        headers = [
            "Name", "DOB", "Gender",
            "Email", "Mobile", "WhatsApp",
            "Address Line 1", "Address Line 2",
            "City", "District", "State", "Country", "Pin Code",
            "Qualification",
            "Marks 10th", "Marks 12th", "Marks UG", "Marks PG",
            "Internship", "Internship Domains",
            
            "Created At"
        ]

        header_font = Font(bold=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font

        # ===== Data Rows =====
        for row_num, student in enumerate(queryset, start=2):
            ws.cell(row=row_num, column=1, value=student.name)
            ws.cell(row=row_num, column=2, value=student.dob)
            ws.cell(row=row_num, column=3, value=student.gender)

            ws.cell(row=row_num, column=4, value=student.email)
            ws.cell(row=row_num, column=5, value=student.mobile)
            ws.cell(row=row_num, column=6, value=student.whatsapp)

            ws.cell(row=row_num, column=7, value=student.address1)
            ws.cell(row=row_num, column=8, value=student.address2)
            ws.cell(row=row_num, column=9, value=student.city)
            ws.cell(row=row_num, column=10, value=student.district)
            ws.cell(row=row_num, column=11, value=student.state)
            ws.cell(row=row_num, column=12, value=student.country)
            ws.cell(row=row_num, column=13, value=student.pin)

            ws.cell(row=row_num, column=14, value=student.qualification)
            ws.cell(row=row_num, column=15, value=student.marks10)
            ws.cell(row=row_num, column=16, value=student.marks12)
            ws.cell(row=row_num, column=17, value=student.marksUG)
            ws.cell(row=row_num, column=18, value=student.marksPG)

            ws.cell(row=row_num, column=19, value=student.internship)
            ws.cell(row=row_num, column=20, value=student.domain)

            # File names (safe)
            # ws.cell(row=row_num, column=21, value=student.aadhar.name if student.aadhar else "")
            # ws.cell(row=row_num, column=22, value=student.collegeid.name if student.collegeid else "")
            # ws.cell(row=row_num, column=23, value=student.cv.name if student.cv else "")
            # ws.cell(row=row_num, column=24, value=student.photo.name if student.photo else "")

            ws.cell(
                row=row_num,
                column=25,
                value=student.created_at.strftime("%d-%m-%Y %H:%M")
                if student.created_at else ""
            )

        # ===== Auto column width =====
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length + 3

        # ===== HTTP Response =====
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=student_membership_full_data.xlsx'

        wb.save(response)
        return response

    download_excel.short_description = "Download selected records as Excel (Full Data)"
